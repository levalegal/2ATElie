"""Export to Excel and PDF."""
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer,
    PageBreak
)


def export_salary_to_excel(records: list, filepath: str, period_start: date, period_end: date):
    """Export salary records to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость"
    ws.append(["Ведомость по зарплате", f"Период: {period_start} — {period_end}"])
    ws.append([])
    headers = ["ФИО", "Должность", "База", "От заказов", "По часам", "Итого"]
    ws.append(headers)
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=6):
        for cell in row:
            cell.font = Font(bold=True)
    for rec in records:
        ws.append([
            rec.get("name", ""),
            rec.get("position", ""),
            rec.get("base", 0),
            rec.get("order_amount", 0),
            rec.get("hourly_amount", 0),
            rec.get("total", 0),
        ])
    wb.save(filepath)


def export_expenses_to_excel(records: list, filepath: str, period_start: date, period_end: date):
    """Export expenses to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Затраты"
    ws.append(["Отчёт по затратам", f"Период: {period_start} — {period_end}"])
    ws.append([])
    headers = ["Дата", "Категория", "Сумма", "Описание"]
    ws.append(headers)
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=4):
        for cell in row:
            cell.font = Font(bold=True)
    for rec in records:
        ws.append([
            rec.get("date", ""),
            rec.get("category", ""),
            rec.get("amount", 0),
            rec.get("description", ""),
        ])
    wb.save(filepath)


def export_salary_to_pdf(records: list, filepath: str, period_start: date, period_end: date):
    """Export salary records to PDF."""
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Ведомость по зарплате", styles["Title"]))
    story.append(Paragraph(f"Период: {period_start} — {period_end}", styles["Normal"]))
    story.append(Spacer(1, 20))
    data = [["ФИО", "Должность", "База", "От заказов", "По часам", "Итого"]]
    for rec in records:
        data.append([
            str(rec.get("name", "")),
            str(rec.get("position", "")),
            str(rec.get("base", 0)),
            str(rec.get("order_amount", 0)),
            str(rec.get("hourly_amount", 0)),
            str(rec.get("total", 0)),
        ])
    t = Table(data)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    story.append(t)
    doc.build(story)


def export_order_invoice(order_data: dict, filepath: str):
    """Export order as invoice PDF."""
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Счёт / Акт", styles["Title"]))
    story.append(Paragraph(f"Заказ №{order_data.get('id', '')}", styles["Normal"]))
    story.append(Paragraph(f"Клиент: {order_data.get('client_name', '')}", styles["Normal"]))
    story.append(Paragraph(f"Дата: {order_data.get('date', '')}", styles["Normal"]))
    story.append(Spacer(1, 20))
    data = [["№", "Наименование", "Кол-во", "Цена", "Сумма"]]
    for i, row in enumerate(order_data.get("materials", []), 1):
        total = row.get("quantity", 0) * row.get("price", 0)
        data.append([str(i), row.get("name", ""), str(row.get("quantity", "")),
                     str(row.get("price", "")), f"{total:,.0f}"])
    data.append(["", "", "", "Итого:", str(order_data.get("total", 0))])
    t = Table(data)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    story.append(t)
    doc.build(story)


def export_employees_to_excel(records: list, filepath: str):
    """Export employees report to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"
    ws.append(["Отчёт по сотрудникам"])
    ws.append([])
    headers = ["ФИО", "Должность", "Телефон", "Ставка", "% от заказа", "₽/ч"]
    ws.append(headers)
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=6):
        for cell in row:
            cell.font = Font(bold=True)
    for rec in records:
        ws.append([
            rec.get("name", ""),
            rec.get("position", ""),
            rec.get("phone", ""),
            rec.get("base_salary", 0),
            rec.get("order_percent", 0),
            rec.get("hourly_rate", 0),
        ])
    wb.save(filepath)


def export_dashboard_to_pdf(stats: dict, filepath: str):
    """Export dashboard stats to PDF."""
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Отчёт по дашборду", styles["Title"]))
    story.append(Paragraph(f"Период: {stats.get('period', '')}", styles["Normal"]))
    story.append(Spacer(1, 20))
    data = [["Показатель", "Значение"]]
    for k, v in stats.get("items", []):
        data.append([k, str(v)])
    t = Table(data)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    story.append(t)
    doc.build(story)


# Atelier brand colors for full export PDF
_ACCENT = colors.HexColor("#8b3a3a")
_HEADER_BG = colors.HexColor("#2d2d30")
_LIGHT_BG = colors.HexColor("#f8f8f8")
_BORDER = colors.HexColor("#3d3d40")


def _styled_table(data, col_widths=None):
    """Create a styled table with Atelier theme."""
    t = Table(data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _ACCENT),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _LIGHT_BG]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d0d0")),
        ("LINEBELOW", (0, 0), (-1, 0), 2, _ACCENT),
    ]))
    return t


def export_all_data_to_pdf(data: dict, filepath: str):
    """
    Export all atelier data to a beautifully formatted PDF.
    data: dict with keys: summary, employees, clients, orders, materials, expenses
    """
    doc = SimpleDocTemplate(
        filepath, pagesize=A4,
        rightMargin=25*mm, leftMargin=25*mm,
        topMargin=20*mm, bottomMargin=20*mm
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "AtelierTitle",
        parent=styles["Title"],
        fontSize=24,
        textColor=_ACCENT,
        spaceAfter=6,
    )
    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontSize=14,
        textColor=_ACCENT,
        spaceBefore=20,
        spaceAfter=10,
    )
    story = []

    # Header block
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    story.append(Paragraph(
        '<font size="28" color="#8b3a3a"><b>Ателье</b></font>',
        styles["Normal"]
    ))
    story.append(Paragraph(
        f'<font size="11" color="#6b6b6b">Полный отчёт · {now}</font>',
        styles["Normal"]
    ))
    story.append(Spacer(1, 24))

    # Summary section
    summary = data.get("summary", {})
    story.append(Paragraph("Сводка", section_style))
    summary_data = [
        ["Показатель", "Значение"],
        ["Выручка", f"{summary.get('revenue', 0):,.0f} ₽"],
        ["Затраты", f"{summary.get('expenses', 0):,.0f} ₽"],
        ["Прибыль", f"{summary.get('profit', 0):,.0f} ₽"],
        ["Заказов всего", str(summary.get("orders_count", 0))],
        ["Сотрудников", str(summary.get("employees_count", 0))],
        ["Клиентов", str(summary.get("clients_count", 0))],
    ]
    story.append(_styled_table(summary_data, [80*mm, 80*mm]))
    story.append(Spacer(1, 16))

    # Employees
    employees = data.get("employees", [])
    if employees:
        story.append(Paragraph("Сотрудники", section_style))
        emp_data = [["ФИО", "Должность", "Телефон", "Ставка", "%", "₽/ч"]]
        for e in employees:
            emp_data.append([
                str(e.get("name", "")),
                str(e.get("position", "")),
                str(e.get("phone", "")),
                f"{e.get('base_salary', 0):,.0f}",
                str(e.get("order_percent", "")),
                f"{e.get('hourly_rate', 0):,.0f}",
            ])
        story.append(_styled_table(emp_data))
        story.append(Spacer(1, 16))

    # Clients
    clients = data.get("clients", [])
    if clients:
        story.append(Paragraph("Клиенты", section_style))
        cli_data = [["ФИО", "Телефон", "Email"]]
        for c in clients:
            cli_data.append([
                str(c.get("full_name", "")),
                str(c.get("phone", "")),
                str(c.get("email", "")),
            ])
        story.append(_styled_table(cli_data))
        story.append(Spacer(1, 16))

    # Orders
    orders = data.get("orders", [])
    if orders:
        story.append(Paragraph("Заказы", section_style))
        ord_data = [["№", "Клиент", "Статус", "Сумма", "Срок"]]
        for o in orders:
            ord_data.append([
                str(o.get("id", "")),
                str(o.get("client_name", ""))[:30],
                str(o.get("status_display", o.get("status", ""))),
                f"{o.get('total_amount', 0):,.0f} ₽",
                str(o.get("deadline", "")),
            ])
        story.append(_styled_table(ord_data))
        story.append(Spacer(1, 16))

    # Materials
    materials = data.get("materials", [])
    if materials:
        story.append(Paragraph("Материалы", section_style))
        mat_data = [["Наименование", "Категория", "Ед.", "Цена", "Остаток"]]
        for m in materials:
            mat_data.append([
                str(m.get("name", "")),
                str(m.get("category", "")),
                str(m.get("unit", "")),
                f"{m.get('price_per_unit', 0):,.0f}",
                str(m.get("quantity", "")),
            ])
        story.append(_styled_table(mat_data))
        story.append(Spacer(1, 16))

    # Expenses
    expenses = data.get("expenses", [])
    if expenses:
        story.append(Paragraph("Затраты", section_style))
        exp_data = [["Дата", "Категория", "Сумма", "Описание"]]
        for e in expenses:
            exp_data.append([
                str(e.get("date", "")),
                str(e.get("category", "")),
                f"{e.get('amount', 0):,.0f} ₽",
                str(e.get("description", ""))[:40],
            ])
        story.append(_styled_table(exp_data))

    # Footer
    story.append(Spacer(1, 30))
    story.append(Paragraph(
        '<font size="9" color="#9a9a9d">— Ателье · Система учёта заказов, зарплат и затрат —</font>',
        styles["Normal"]
    ))

    doc.build(story)
